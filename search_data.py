from tkinter import messagebox

from openpyxl import load_workbook


def find_in_excel_table(file_path: object, sheet_name: object, table_name: object, col_name: object, search_value: object,
                        data_gas_num, data_light_num, data_water_num) -> str:
    print('>>>>>>>',search_value)
    wb = load_workbook(file_path)
    # ws = wb[sheet_name]
    ws = wb.active

    # Знаходимо таблицю за ім'ям
    if table_name not in ws.tables:
        return messagebox.showinfo("Не вірне значення", "Таблицю не знайдено")

    table = ws.tables[table_name]
    table_range = ws[table.ref]  # Отримуємо діапазон клітинок (напр., A1:C10)

    # Визначаємо індекс потрібної колонки
    header_row = table_range[0]
    col_idx = None
    for i, cell in enumerate(header_row):
        if cell.value == col_name:
            col_idx = i
            break

    if col_idx is None:
        return messagebox.showinfo("Не вірне значення", "Колонку не знайдено. Введіть вірно назву колонки.")

    # Пошук значення в обраній колонці (пропускаючи заголовок)
    for row in table_range[1:]:
        if row[col_idx].value == search_value:
            r = row[col_idx].row
            print(r)
            ws[f'C{r}'] = data_gas_num
            ws[f'J{r}'] = data_light_num
            ws[f'P{r}'] = data_water_num
            wb.save(file_path)
            messagebox.showinfo("Інформація", "Данні додані до файлу")

            return f"Знайдено у рядку {row[col_idx].row}"

    return messagebox.showinfo("Не вірно значення", "Значення не знайдено. Введіть вірно місяць.")



# Приклад використання
# result = find_in_excel_table('k_30.xlsx', 'A1', 'year2026', 'Дата', 'квітень') # добавити 3 параметра
# print(result)